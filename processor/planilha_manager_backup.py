#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📁 ARQUIVO: processor/planilha_manager.py
💾 FUNÇÃO: Gerenciador de planilhas ENEL incremental
🔧 DESCRIÇÃO: Sistema incremental de controle ENEL - baseado no BRK
👨‍💼 AUTOR: Baseado na lógica do sistema desktop ENEL
🔒 ESTRUTURA: Planilha relacionamento + controle mensal incremental
"""

import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import requests
import base64

# Imports condicionais - Remover dependência do pandas
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# pandas opcional (apenas fallback para ambiente local)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

class PlanilhaManagerEnel:
    """
    Gerenciador de planilhas ENEL incremental (VERSÃO SEM PANDAS - RENDER COMPATÍVEL)
    
    Funcionalidades:
    - Carregar planilha de relacionamento SEM pandas
    - Criar/atualizar planilha de controle mensal SEM pandas
    - Processar faturas incrementalmente
    - Detectar instalações faltantes
    - Gerar relatórios de status
    
    BASEADO NO PADRÃO BRK que funciona perfeitamente no Render
    """
    
    def __init__(self, auth_manager, onedrive_manager):
        """
        Inicializar gerenciador de planilhas ENEL
        
        Args:
            auth_manager: Instância MicrosoftAuth
            onedrive_manager: Instância OneDriveManagerEnel
        """
        self.auth = auth_manager
        self.onedrive = onedrive_manager
        
        # Dados das planilhas (versão sem pandas - Render compatible)
        self.relacionamentos_dados = []  # List[Dict] - sem pandas
        self.controle_mensal_dados = []  # List[Dict] - sem pandas
        
        # Cache das planilhas (fallback pandas para ambiente local)
        self.planilha_relacionamento = None
        self.planilha_controle_atual = None
        self.mes_atual = datetime.now().month
        self.ano_atual = datetime.now().year
        
        # Configurações
        self.nome_planilha_relacionamento = "instalacao_enel.xlsx"
        self.colunas_relacionamento = {
            'casa_oracao': None,      # Será detectada automaticamente  
            'instalacao': None,       # Será detectada automaticamente
            'dia_vencimento': None    # Será detectada automaticamente
        }
        
        print(f"📊 Planilha Manager ENEL inicializado")
        print(f"   Período atual: {self.mes_atual:02d}/{self.ano_atual}")
    
    def carregar_planilha_relacionamento(self) -> bool:
        """
        Carregar planilha de relacionamento ENEL do OneDrive SEM PANDAS
        
        Estrutura esperada:
        A: Casa de Oração
        B: Número Instalação  
        C: Dia Vencimento
        
        Returns:
            bool: True se carregada com sucesso
        """
        # Tentar versão sem pandas primeiro (Render compatible)
        resultado_sem_pandas = self.carregar_planilha_relacionamento_sem_pandas()
        if resultado_sem_pandas:
            return True
        
        # Fallback para versão com pandas (ambiente local)
        if PANDAS_AVAILABLE:
            return self._carregar_planilha_relacionamento_com_pandas()
        else:
            print("⚠️ Não foi possível carregar planilha (pandas não disponível)")
            return False
            
            if not self.onedrive.pasta_planilhas_id:
                print("❌ Pasta planilhas OneDrive não configurada")
                return False
            
            # Buscar arquivo no OneDrive
            planilha_bytes = self._baixar_arquivo_onedrive(
                self.onedrive.pasta_planilhas_id, 
                self.nome_planilha_relacionamento
            )
            
            if not planilha_bytes:
                print(f"⚠️ Planilha {self.nome_planilha_relacionamento} não encontrada")
                return self._criar_planilha_relacionamento_exemplo()
            
            # Este método foi refatorado - ver carregar_planilha_relacionamento()
            
            # Detectar colunas automaticamente
            if not self._detectar_colunas_relacionamento():
                print("❌ Não foi possível detectar colunas da planilha relacionamento")
                return False
            
            print(f"✅ Planilha relacionamento carregada: {len(self.planilha_relacionamento)} instalações")
            return True
            
        except Exception as e:
            print(f"❌ Erro carregando planilha relacionamento: {e}")
            return False
    
    def _detectar_colunas_relacionamento(self) -> bool:
        """Detectar colunas da planilha relacionamento automaticamente"""
        try:
            if self.planilha_relacionamento is None:
                return False
            
            # Limpar nomes das colunas
            self.planilha_relacionamento.columns = self.planilha_relacionamento.columns.str.strip()
            
            # Detectar coluna Casa de Oração
            for col in self.planilha_relacionamento.columns:
                col_lower = col.lower()
                if any(palavra in col_lower for palavra in ['casa', 'oração', 'oracao', 'nome']):
                    self.colunas_relacionamento['casa_oracao'] = col
                    break
            
            # Detectar coluna Instalação
            for col in self.planilha_relacionamento.columns:
                col_lower = col.lower()
                if any(palavra in col_lower for palavra in ['instalação', 'instalacao', 'enel', 'numero']):
                    self.colunas_relacionamento['instalacao'] = col
                    break
            
            # Detectar coluna Dia Vencimento
            for col in self.planilha_relacionamento.columns:
                col_lower = col.lower()
                if any(palavra in col_lower for palavra in ['vencimento', 'vence', 'dia']):
                    self.colunas_relacionamento['dia_vencimento'] = col
                    break
            
            # Verificar se todas foram encontradas
            colunas_ok = all(v is not None for v in self.colunas_relacionamento.values())
            
            if colunas_ok:
                print(f"🎯 Colunas detectadas:")
                print(f"   Casa: '{self.colunas_relacionamento['casa_oracao']}'")
                print(f"   Instalação: '{self.colunas_relacionamento['instalacao']}'")
                print(f"   Vencimento: '{self.colunas_relacionamento['dia_vencimento']}'")
            
            return colunas_ok
            
        except Exception as e:
            print(f"❌ Erro detectando colunas: {e}")
            return False
    
    def _criar_planilha_relacionamento_exemplo(self) -> bool:
        """Criar planilha de relacionamento exemplo se não existir"""
        try:
            # Dados exemplo
            dados_exemplo = {
                'Casa de Oração': [
                    'PIA Central',
                    'PIA Norte', 
                    'PIA Sul',
                    'Matriz',
                    'Casa Exemplo'
                ],
                'Numero Instalacao': [
                    '123456789',
                    '987654321',
                    '555666777',
                    '111222333',
                    '999888777'
                ],
                'Dia Vencimento': [15, 20, 25, 10, 5]
            }
            
            self.planilha_relacionamento = pd.DataFrame(dados_exemplo)
            
            # Configurar colunas
            self.colunas_relacionamento = {
                'casa_oracao': 'Casa de Oração',
                'instalacao': 'Numero Instalacao', 
                'dia_vencimento': 'Dia Vencimento'
            }
            
            # Salvar no OneDrive
            if self._salvar_planilha_relacionamento():
                print("✅ Planilha relacionamento exemplo criada")
                return True
            else:
                print("⚠️ Erro salvando planilha exemplo")
                return False
                
        except Exception as e:
            print(f"❌ Erro criando planilha exemplo: {e}")
            return False
    
    def carregar_planilha_controle_mensal(self, mes: int = None, ano: int = None) -> bool:
        """
        Carregar ou criar planilha de controle mensal
        
        Args:
            mes (int): Mês (default: atual)
            ano (int): Ano (default: atual)
            
        Returns:
            bool: True se carregada/criada com sucesso
        """
        try:
            if not EXCEL_AVAILABLE or not self.planilha_relacionamento is not None:
                return False
            
            mes = mes or self.mes_atual
            ano = ano or self.ano_atual
            
            nome_controle = f"controle_enel_{ano}_{mes:02d}.xlsx"
            
            # Tentar carregar existente
            controle_bytes = self._baixar_arquivo_onedrive(
                self.onedrive.pasta_planilhas_id,
                nome_controle
            )
            
            if controle_bytes:
                # Carregar existente
                self.planilha_controle_atual = pd.read_excel(controle_bytes)
                print(f"✅ Controle mensal carregado: {nome_controle}")
            else:
                # Criar novo baseado no relacionamento
                self.planilha_controle_atual = self._criar_controle_mensal(mes, ano)
                print(f"📊 Controle mensal criado: {nome_controle}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro carregando controle mensal: {e}")
            return False
    
    def _criar_controle_mensal(self, mes: int, ano: int) -> pd.DataFrame:
        """
        Criar planilha de controle mensal baseada no relacionamento
        
        Args:
            mes (int): Mês da competência
            ano (int): Ano da competência
            
        Returns:
            pd.DataFrame: Planilha de controle inicializada
        """
        try:
            # Base da planilha relacionamento
            controle = self.planilha_relacionamento.copy()
            
            # Renomear colunas para padrão
            mapeamento_colunas = {}
            for coluna_tipo, nome_coluna in self.colunas_relacionamento.items():
                if coluna_tipo == 'casa_oracao':
                    mapeamento_colunas[nome_coluna] = 'Casa de Oração'
                elif coluna_tipo == 'instalacao':
                    mapeamento_colunas[nome_coluna] = 'Numero_Instalacao'
                elif coluna_tipo == 'dia_vencimento':
                    mapeamento_colunas[nome_coluna] = 'Dia_Vencimento'
            
            controle = controle.rename(columns=mapeamento_colunas)
            
            # Adicionar colunas de controle (igual ao desktop ENEL)
            controle['Status'] = 'Faltando'
            controle['Data_Vencimento'] = pd.NaType
            controle['Valor_Total'] = 0.0
            controle['Consumo_kWh'] = 0
            controle['Energia_Injetada'] = 0.0
            controle['Energia_Compensada'] = 0.0
            controle['Valor_Sem_ICMS'] = 0.0
            controle['TUSD'] = 0.0
            controle['TE'] = 0.0
            controle['Arquivo_PDF'] = ''
            controle['Data_Processamento'] = pd.NaType
            controle['Competencia'] = f"{mes:02d}/{ano}"
            
            # Calcular data de vencimento esperada
            for idx, row in controle.iterrows():
                try:
                    dia = int(row['Dia_Vencimento'])
                    data_venc = datetime(ano, mes, dia).strftime('%d/%m/%Y')
                    controle.loc[idx, 'Data_Vencimento'] = data_venc
                except:
                    controle.loc[idx, 'Data_Vencimento'] = f"15/{mes:02d}/{ano}"
            
            return controle
            
        except Exception as e:
            print(f"❌ Erro criando controle mensal: {e}")
            return pd.DataFrame()
    
    def processar_fatura_incremental(self, dados_fatura: Dict) -> bool:
        """
        Processar uma fatura incrementalmente na planilha de controle
        
        Args:
            dados_fatura (Dict): Dados extraídos da fatura PDF
            
        Returns:
            bool: True se processada com sucesso
        """
        try:
            if self.planilha_controle_atual is None:
                print("❌ Planilha controle não carregada")
                return False
            
            numero_instalacao = str(dados_fatura.get('numero_instalacao', '')).strip()
            
            # Buscar linha correspondente na planilha
            mask = self.planilha_controle_atual['Numero_Instalacao'].astype(str).str.strip() == numero_instalacao
            linha_idx = self.planilha_controle_atual.index[mask].tolist()
            
            if not linha_idx:
                print(f"⚠️ Instalação {numero_instalacao} não encontrada na planilha relacionamento")
                return False
            
            idx = linha_idx[0]
            
            # Atualizar dados na planilha (mesmos campos do desktop)
            self.planilha_controle_atual.loc[idx, 'Status'] = 'Recebida'
            self.planilha_controle_atual.loc[idx, 'Valor_Total'] = dados_fatura.get('valor_total_num', 0.0)
            self.planilha_controle_atual.loc[idx, 'Consumo_kWh'] = dados_fatura.get('consumo_kwh_num', 0)
            self.planilha_controle_atual.loc[idx, 'Energia_Injetada'] = dados_fatura.get('energia_injetada', 0.0)
            self.planilha_controle_atual.loc[idx, 'Energia_Compensada'] = dados_fatura.get('energia_compensada', 0.0)
            self.planilha_controle_atual.loc[idx, 'Valor_Sem_ICMS'] = dados_fatura.get('valor_sem_icms', 0.0)
            self.planilha_controle_atual.loc[idx, 'TUSD'] = dados_fatura.get('tusd', 0.0)
            self.planilha_controle_atual.loc[idx, 'TE'] = dados_fatura.get('te', 0.0)
            self.planilha_controle_atual.loc[idx, 'Arquivo_PDF'] = dados_fatura.get('arquivo', '')
            self.planilha_controle_atual.loc[idx, 'Data_Processamento'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            
            # Data de vencimento se disponível
            if dados_fatura.get('data_vencimento'):
                self.planilha_controle_atual.loc[idx, 'Data_Vencimento'] = dados_fatura['data_vencimento']
            
            casa_oracao = self.planilha_controle_atual.loc[idx, 'Casa de Oração']
            print(f"✅ Fatura processada: {casa_oracao} - R$ {dados_fatura.get('valor_total_num', 0.0):.2f}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro processando fatura: {e}")
            return False
    
    def obter_status_controle(self) -> Dict:
        """
        Obter status atual do controle mensal
        
        Returns:
            Dict: Status detalhado
        """
        try:
            if self.planilha_controle_atual is None:
                return {"erro": "Planilha controle não carregada"}
            
            total_instalacoes = len(self.planilha_controle_atual)
            recebidas = len(self.planilha_controle_atual[self.planilha_controle_atual['Status'] == 'Recebida'])
            faltando = total_instalacoes - recebidas
            
            # Valor total processado
            valor_total = self.planilha_controle_atual[
                self.planilha_controle_atual['Status'] == 'Recebida'
            ]['Valor_Total'].sum()
            
            # Instalações faltantes
            faltantes = self.planilha_controle_atual[
                self.planilha_controle_atual['Status'] == 'Faltando'
            ][['Casa de Oração', 'Numero_Instalacao', 'Data_Vencimento']].to_dict('records')
            
            return {
                "competencia": f"{self.mes_atual:02d}/{self.ano_atual}",
                "total_instalacoes": total_instalacoes,
                "faturas_recebidas": recebidas,
                "faturas_faltando": faltando,
                "percentual_completo": round((recebidas / total_instalacoes) * 100, 1),
                "valor_total_processado": round(valor_total, 2),
                "instalacoes_faltantes": faltantes,
                "timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            return {"erro": str(e)}
    
    def salvar_controle_mensal(self) -> bool:
        """Salvar planilha de controle atual no OneDrive"""
        try:
            if self.planilha_controle_atual is None:
                return False
            
            nome_arquivo = f"controle_enel_{self.ano_atual}_{self.mes_atual:02d}.xlsx"
            
            # Converter para bytes
            from io import BytesIO
            buffer = BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                self.planilha_controle_atual.to_excel(writer, sheet_name='Controle ENEL', index=False)
                
                # Aplicar formatação básica
                workbook = writer.book
                worksheet = writer.sheets['Controle ENEL']
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            buffer.seek(0)
            arquivo_bytes = buffer.getvalue()
            
            # Upload para OneDrive
            sucesso = self.onedrive.upload_arquivo(
                arquivo_bytes,
                nome_arquivo,
                self.onedrive.pasta_planilhas_id
            )
            
            if sucesso:
                print(f"💾 Controle mensal salvo: {nome_arquivo}")
                return True
            else:
                print(f"❌ Erro salvando controle mensal")
                return False
                
        except Exception as e:
            print(f"❌ Erro salvando controle: {e}")
            return False
    
    def _salvar_planilha_relacionamento(self) -> bool:
        """Salvar planilha de relacionamento no OneDrive"""
        try:
            if self.planilha_relacionamento is None:
                return False
            
            from io import BytesIO
            buffer = BytesIO()
            self.planilha_relacionamento.to_excel(buffer, index=False)
            buffer.seek(0)
            
            return self.onedrive.upload_arquivo(
                buffer.getvalue(),
                self.nome_planilha_relacionamento,
                self.onedrive.pasta_planilhas_id
            )
            
        except Exception as e:
            print(f"❌ Erro salvando planilha relacionamento: {e}")
            return False
    
    def carregar_planilha_relacionamento_sem_pandas(self) -> bool:
        """
        Carregar planilha de relacionamento ENEL SEM pandas (Render compatible)
        Baseado no padrão BRK que funciona perfeitamente
        
        Returns:
            bool: True se carregada com sucesso
        """
        try:
            if not EXCEL_AVAILABLE:
                print("⚠️ openpyxl não disponível")
                return False
            
            if not self.onedrive.pasta_enel_id:
                print("❌ Pasta raiz ENEL OneDrive não configurada")
                return False
            
            print(f"📁 Carregando planilha {self.nome_planilha_relacionamento} da raiz ENEL (igual BRK)...")
            
            # Buscar arquivo na pasta raiz ENEL (igual BRK)
            planilha_bytes = self._baixar_arquivo_onedrive(
                self.onedrive.pasta_enel_id, 
                self.nome_planilha_relacionamento
            )
            
            if not planilha_bytes:
                print(f"⚠️ Planilha {self.nome_planilha_relacionamento} não encontrada")
                return self._criar_planilha_relacionamento_exemplo_sem_pandas()
            
            # Processar Excel manualmente (sem pandas)
            registros = self._processar_excel_relacionamento_manual(planilha_bytes)
            
            if not registros:
                print(f"❌ Nenhum registro encontrado na planilha")
                return False
            
            # Converter para estrutura interna (sem pandas)
            self.relacionamentos_dados = registros
            
            print(f"✅ PLANILHA DE RELACIONAMENTO CARREGADA SEM pandas!")
            print(f"   📄 Registros: {len(registros)}")
            print(f"   🏠 Casas: {len(set(r.get('Casa', '') for r in registros))}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro carregando planilha sem pandas: {e}")
            return False
    
    def _processar_excel_relacionamento_manual(self, excel_bytes) -> List[Dict]:
        """
        Processar Excel de relacionamento manualmente (sem pandas)
        Estrutura: A=Casa, B=Instalação, C=Vencimento
        
        Args:
            excel_bytes: Bytes do arquivo Excel
            
        Returns:
            List[Dict]: Registros processados
        """
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            import io
            
            registros = []
            
            # Excel (.xlsx) é um arquivo ZIP com XMLs internos
            with zipfile.ZipFile(io.BytesIO(excel_bytes), 'r') as zip_file:
                
                # 1. Ler shared strings (textos da planilha)
                shared_strings = []
                try:
                    with zip_file.open('xl/sharedStrings.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        
                        # Namespace do Excel
                        ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                        for si in root.findall('.//ss:si', ns):
                            t_elem = si.find('ss:t', ns)
                            if t_elem is not None:
                                shared_strings.append(t_elem.text or '')
                            else:
                                shared_strings.append('')
                except:
                    pass  # Planilha pode não ter shared strings
                
                # 2. Ler dados da primeira worksheet
                try:
                    with zip_file.open('xl/worksheets/sheet1.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        
                        ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                        # Processar linhas (pular cabeçalho)
                        rows = root.findall('.//ss:row', ns)
                        for row_idx, row in enumerate(rows[1:], 2):  # Pular linha 1 (cabeçalho)
                            cells = row.findall('ss:c', ns)
                            
                            casa = ''
                            instalacao = ''
                            vencimento = ''
                            
                            for cell in cells:
                                cell_ref = cell.get('r', '')
                                col_letter = ''.join([c for c in cell_ref if c.isalpha()])
                                
                                # Obter valor da célula
                                valor = ''
                                v_elem = cell.find('ss:v', ns)
                                if v_elem is not None:
                                    if cell.get('t') == 's':  # String compartilhada
                                        try:
                                            idx = int(v_elem.text)
                                            if 0 <= idx < len(shared_strings):
                                                valor = shared_strings[idx]
                                        except (ValueError, IndexError):
                                            valor = v_elem.text or ''
                                    else:
                                        valor = v_elem.text or ''
                                
                                # Mapear colunas
                                if col_letter == 'A':
                                    casa = str(valor).strip()
                                elif col_letter == 'B':
                                    instalacao = str(valor).strip()
                                elif col_letter == 'C':
                                    vencimento = str(valor).strip()
                            
                            # Adicionar registro se válido
                            if casa and instalacao:
                                registros.append({
                                    'Casa': casa,
                                    'Instalacao': instalacao,
                                    'Vencimento': vencimento or '0'
                                })
                
                except Exception as e:
                    print(f"⚠️ Erro processando worksheet: {e}")
            
            return registros
            
        except Exception as e:
            print(f"❌ Erro processando Excel manual: {e}")
            return []
    
    def _criar_planilha_relacionamento_exemplo_sem_pandas(self) -> bool:
        """
        Criar planilha de relacionamento exemplo SEM pandas
        """
        try:
            if not EXCEL_AVAILABLE:
                return False
            
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relacionamento ENEL"
            
            # Cabeçalhos
            ws.cell(row=1, column=1, value="Casa de Oração")
            ws.cell(row=1, column=2, value="Número Instalação")
            ws.cell(row=1, column=3, value="Dia Vencimento")
            
            # Dados de exemplo
            exemplos = [
                ("Casa Exemplo 1", "12345678", "10"),
                ("Casa Exemplo 2", "87654321", "15"),
                ("Casa Exemplo 3", "11223344", "20")
            ]
            
            for row_idx, (casa, instalacao, vencimento) in enumerate(exemplos, 2):
                ws.cell(row=row_idx, column=1, value=casa)
                ws.cell(row=row_idx, column=2, value=instalacao)
                ws.cell(row=row_idx, column=3, value=vencimento)
            
            # Salvar no OneDrive
            buffer = io.BytesIO()
            wb.save(buffer)
            planilha_bytes = buffer.getvalue()
            buffer.close()
            
            sucesso = self._upload_arquivo_onedrive(
                self.onedrive.pasta_planilhas_id,
                self.nome_planilha_relacionamento,
                planilha_bytes
            )
            
            if sucesso:
                # Definir dados exemplo na memória
                self.relacionamentos_dados = [
                    {'Casa': casa, 'Instalacao': instalacao, 'Vencimento': vencimento}
                    for casa, instalacao, vencimento in exemplos
                ]
                print(f"✅ Planilha exemplo criada SEM pandas!")
                return True
            
            return False
            
        except Exception as e:
            print(f"❌ Erro criando exemplo sem pandas: {e}")
            return False
    
    def _carregar_planilha_relacionamento_com_pandas(self) -> bool:
        """
        Versão original com pandas (fallback para ambiente local)
        """
        if not PANDAS_AVAILABLE:
            return False
            
        try:
            if not self.onedrive.pasta_planilhas_id:
                print("❌ Pasta planilhas OneDrive não configurada")
                return False
            
            # Buscar arquivo no OneDrive
            planilha_bytes = self._baixar_arquivo_onedrive(
                self.onedrive.pasta_planilhas_id, 
                self.nome_planilha_relacionamento
            )
            
            if not planilha_bytes:
                print(f"⚠️ Planilha {self.nome_planilha_relacionamento} não encontrada")
                return self._criar_planilha_relacionamento_exemplo_com_pandas()
            
            # Carregar em pandas
            self.planilha_relacionamento = pd.read_excel(planilha_bytes)
            
            print(f"✅ Planilha relacionamento carregada COM pandas: {len(self.planilha_relacionamento)} registros")
            return True
            
        except Exception as e:
            print(f"❌ Erro carregando com pandas: {e}")
            return False
    
    def _baixar_arquivo_onedrive(self, pasta_id: str, nome_arquivo: str) -> Optional[bytes]:
        """Baixar arquivo do OneDrive"""
        try:
            if not self.auth.access_token:
                return None
            
            headers = self.auth.obter_headers_autenticados()
            
            # Buscar arquivo na pasta
            url = f"https://graph.microsoft.com/v1.0/me/drive/items/{pasta_id}/children"
            params = {"$filter": f"name eq '{nome_arquivo}'"}
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                items = response.json().get('value', [])
                
                if items:
                    file_id = items[0]['id']
                    
                    # Baixar conteúdo
                    content_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/content"
                    content_response = requests.get(content_url, headers=headers, timeout=60)
                    
                    if content_response.status_code == 200:
                        return content_response.content
            
            return None
            
        except Exception as e:
            print(f"❌ Erro baixando {nome_arquivo}: {e}")
            return None