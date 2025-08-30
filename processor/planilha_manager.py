#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📁 ARQUIVO: processor/planilha_manager.py
💾 FUNÇÃO: Gerenciador de planilhas ENEL incremental
🔧 DESCRIÇÃO: Sistema incremental de controle ENEL - baseado no BRK
👨‍💼 AUTOR: Baseado na lógica do sistema desktop ENEL
🔒 ESTRUTURA: Planilha relacionamento + controle mensal incremental
"""

import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import requests
import base64

# Imports condicionais - Remover dependência do pandas
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# pandas removido - Render não suporta
# try:
#     import pandas as pd
#     PANDAS_AVAILABLE = True
# except ImportError:
PANDAS_AVAILABLE = False

# Import cálculo centralizado
from .calculo_enel import calcular_media_e_diferenca_enel

class PlanilhaManagerEnel:
    """
    Gerenciador de planilhas ENEL incremental (VERSÃO SEM PANDAS - RENDER COMPATÍVEL)
    
    Funcionalidades:
    - Carregar planilha de relacionamento SEM pandas
    - Criar/atualizar planilha de controle mensal SEM pandas
    - Processar faturas incrementalmente
    - Detectar instalações faltantes
    - Gerar relatórios de status
    
    BASEADO NO PADRÃO BRK que funciona perfeitamente no Render
    """
    
    def __init__(self, auth_manager, onedrive_manager, database_manager=None):
        """
        Inicializar gerenciador de planilhas ENEL
        
        Args:
            auth_manager: Instância MicrosoftAuth
            onedrive_manager: Instância OneDriveManagerEnel
            database_manager: Instância DatabaseEnel (opcional)
        """
        self.auth = auth_manager
        self.onedrive = onedrive_manager
        self.database = database_manager
        
        # Dados das planilhas (versão sem pandas - Render compatible)
        self.relacionamentos_dados = []  # List[Dict] - sem pandas
        self.controle_mensal_dados = []  # List[Dict] - sem pandas
        
        # Cache das planilhas OneDrive ENEL
        self.planilha_relacionamento = None
        self.planilha_controle_atual = None
        self.mes_atual = datetime.now().month
        self.ano_atual = datetime.now().year
        
        # Estatísticas de duplicatas (igual BRK)
        self.estatisticas_duplicatas = {
            "emails_duplicados": 0,
            "faturas_duplicadas": 0,
            "instalacoes_reprocessadas": [],
            "timestamp_processamento": None
        }
        
        # Configurações
        self.nome_planilha_relacionamento = "relação_enel.xlsx"
        self.colunas_relacionamento = {
            'casa_oracao': 'Casa de Oração',     # Coluna A (com código BRK!)
            'instalacao': 'Instalação Enel',     # Coluna B  
            'dia_vencimento': 'Vencimento'       # Coluna C
        }
        
        print(f"📊 Planilha Manager ENEL inicializado")
        print(f"   Período atual: {self.mes_atual:02d}/{self.ano_atual}")
    
    def carregar_planilha_relacionamento(self) -> bool:
        """
        Carregar planilha de relacionamento ENEL do OneDrive SEM PANDAS
        
        Estrutura esperada:
        A: Casa de Oração
        B: Número Instalação  
        C: Dia Vencimento
        
        Returns:
            bool: True se carregada com sucesso
        """
        # Tentar versão sem pandas primeiro (Render compatible)
        resultado_sem_pandas = self.carregar_planilha_relacionamento_sem_pandas()
        if resultado_sem_pandas:
            return True
        
        # Fallback para versão com pandas (ambiente local) 
        if PANDAS_AVAILABLE:
            return self._carregar_planilha_relacionamento_com_pandas()
        else:
            print("⚠️ Planilha não encontrada - usando modo SEM pandas")
            return False
    
    def _detectar_colunas_relacionamento(self) -> bool:
        """Detectar colunas da planilha relacionamento automaticamente"""
        try:
            if self.planilha_relacionamento is None:
                return False
            
            # Limpar nomes das colunas
            self.planilha_relacionamento.columns = self.planilha_relacionamento.columns.str.strip()
            
            # Detectar coluna Casa de Oração
            for col in self.planilha_relacionamento.columns:
                col_lower = col.lower()
                if any(palavra in col_lower for palavra in ['casa', 'oração', 'oracao', 'nome']):
                    self.colunas_relacionamento['casa_oracao'] = col
                    break
            
            # Detectar coluna Instalação
            for col in self.planilha_relacionamento.columns:
                col_lower = col.lower()
                if any(palavra in col_lower for palavra in ['instalação', 'instalacao', 'enel', 'numero']):
                    self.colunas_relacionamento['instalacao'] = col
                    break
            
            # Detectar coluna Dia Vencimento
            for col in self.planilha_relacionamento.columns:
                col_lower = col.lower()
                if any(palavra in col_lower for palavra in ['vencimento', 'vence', 'dia']):
                    self.colunas_relacionamento['dia_vencimento'] = col
                    break
            
            # Verificar se todas foram encontradas
            colunas_ok = all(v is not None for v in self.colunas_relacionamento.values())
            
            if colunas_ok:
                print(f"🎯 Colunas detectadas:")
                print(f"   Casa: '{self.colunas_relacionamento['casa_oracao']}'")
                print(f"   Instalação: '{self.colunas_relacionamento['instalacao']}'")
                print(f"   Vencimento: '{self.colunas_relacionamento['dia_vencimento']}'")
            
            return colunas_ok
            
        except Exception as e:
            print(f"❌ Erro detectando colunas: {e}")
            return False
    
    def _criar_planilha_relacionamento_exemplo(self) -> bool:
        """Criar planilha de relacionamento exemplo se não existir"""
        try:
            # Dados exemplo
            dados_exemplo = {
                'Casa de Oração': [
                    'PIA Central',
                    'PIA Norte', 
                    'PIA Sul',
                    'Matriz',
                    'Casa Exemplo'
                ],
                'Numero Instalacao': [
                    '123456789',
                    '987654321',
                    '555666777',
                    '111222333',
                    '999888777'
                ],
                'Dia Vencimento': [15, 20, 25, 10, 5]
            }
            
            self.planilha_relacionamento = pd.DataFrame(dados_exemplo)
            
            # Configurar colunas
            self.colunas_relacionamento = {
                'casa_oracao': 'Casa de Oração',
                'instalacao': 'Numero Instalacao', 
                'dia_vencimento': 'Dia Vencimento'
            }
            
            # Salvar no OneDrive
            if self._salvar_planilha_relacionamento():
                print("✅ Planilha relacionamento exemplo criada")
                return True
            else:
                print("⚠️ Erro salvando planilha exemplo")
                return False
                
        except Exception as e:
            print(f"❌ Erro criando planilha exemplo: {e}")
            return False
    
    def carregar_planilha_controle_mensal(self, mes: int = None, ano: int = None) -> bool:
        """
        Carregar ou criar planilha de controle mensal (VERSÃO SEM PANDAS)
        
        Args:
            mes (int): Mês (default: atual)
            ano (int): Ano (default: atual)
            
        Returns:
            bool: True se carregada/criada com sucesso
        """
        try:
            mes = mes or self.mes_atual
            ano = ano or self.ano_atual
            
            nome_controle = f"controle_enel_{ano}_{mes:02d}.xlsx"
            
            # Se não temos relacionamentos carregados, carregar primeiro
            if not self.relacionamentos_dados:
                if not self.carregar_planilha_relacionamento():
                    print("❌ Erro: Planilha relacionamento necessária para criar controle")
                    return False
            
            # Tentar carregar existente (não implementado ainda)
            # controle_bytes = self._baixar_arquivo_onedrive(self.onedrive.pasta_planilhas_id, nome_controle)
            
            # Por enquanto sempre criar novo baseado no relacionamento
            self.controle_mensal_dados = self._criar_controle_mensal_sem_pandas(mes, ano)
            
            if self.controle_mensal_dados:
                print(f"📊 Controle mensal criado SEM pandas: {nome_controle}")
                print(f"   📄 Registros: {len(self.controle_mensal_dados)}")
                return True
            else:
                print(f"❌ Erro criando controle mensal")
                return False
            
        except Exception as e:
            print(f"❌ Erro carregando controle mensal: {e}")
            return False
    
    def _criar_controle_mensal_sem_pandas(self, mes: int, ano: int) -> List[Dict]:
        """
        Criar planilha de controle mensal baseada no relacionamento (SEM PANDAS - Render compatible)
        
        Args:
            mes (int): Mês da competência
            ano (int): Ano da competência
            
        Returns:
            List[Dict]: Dados da planilha de controle inicializada
        """
        try:
            controle_dados = []
            
            # Usar relacionamentos carregados sem pandas
            for relacionamento in self.relacionamentos_dados:
                casa_oracao = relacionamento.get('Casa', '')
                numero_instalacao = relacionamento.get('Instalacao', '')
                dia_vencimento = relacionamento.get('Vencimento', '15')
                
                # Calcular data de vencimento esperada
                try:
                    dia = int(dia_vencimento) if dia_vencimento.isdigit() else 15
                    data_venc = datetime(ano, mes, dia).strftime('%d/%m/%Y')
                except:
                    data_venc = f"15/{mes:02d}/{ano}"
                
                # 18 colunas EXATAS conforme planilha real ENEL
                registro = {
                    # 1. Casa de Oração
                    'Casa de Oração': casa_oracao,
                    # 2. Competencia
                    'Competencia': f"{mes:02d}/{ano}",
                    # 3. Data_Emissao
                    'Data_Emissao': '',
                    # 4. Nota_Fiscal
                    'Nota_Fiscal': '',
                    # 5. Vencimento
                    'Vencimento': data_venc,
                    # 6. Valor
                    'Valor': '',
                    # 7. Consumo_kWh
                    'Consumo_kWh': '',
                    # 8. Media_6_Meses
                    'Media_6_Meses': 0.0,
                    # 9. Diferenca_Percentual
                    'Diferenca_Percentual': 0.0,
                    # 10. Porcentagem_Consumo
                    'Porcentagem_Consumo': 100.0,
                    # 11. Alerta_Consumo
                    'Alerta_Consumo': 'Normal',
                    # 12. Sistema_Fotovoltaico
                    'Sistema_Fotovoltaico': 'Não',
                    # 13. Compensacao_TUSD
                    'Compensacao_TUSD': 0.0,
                    # 14. Compensacao_TE
                    'Compensacao_TE': 0.0,
                    # 15. Total_Compensacao
                    'Total_Compensacao': 0.0,
                    # 16. Valor_Integral_Sem_FV
                    'Valor_Integral_Sem_FV': 0.0,
                    # 17. Percentual_Economia_FV
                    'Percentual_Economia_FV': 0.0,
                    # 18. Numero_Instalacao
                    'Numero_Instalacao': numero_instalacao,
                    
                    # Status para controle de duplicatas (padrão BRK)
                    'Status': 'Faltando'
                }
                
                controle_dados.append(registro)
            
            return controle_dados
            
        except Exception as e:
            print(f"❌ Erro criando controle mensal sem pandas: {e}")
            return []
    
    def processar_fatura_incremental(self, dados_fatura: Dict) -> bool:
        """
        Processar uma fatura incrementalmente (método principal - chama versão sem pandas)
        
        Args:
            dados_fatura (Dict): Dados extraídos da fatura PDF
            
        Returns:
            bool: True se processada com sucesso
        """
        return self.processar_fatura_incremental_sem_pandas(dados_fatura)
    
    def processar_fatura_incremental_sem_pandas(self, dados_fatura: Dict) -> bool:
        """
        Processar uma fatura incrementalmente na planilha de controle (SEM PANDAS)
        COM CONTROLE DE DUPLICATAS (baseado no padrão BRK)
        
        Args:
            dados_fatura (Dict): Dados extraídos da fatura PDF
            
        Returns:
            bool: True se processada com sucesso
        """
        try:
            if not self.controle_mensal_dados:
                print("❌ Controle mensal não carregado")
                return False
            
            numero_instalacao = str(dados_fatura.get('numero_instalacao', '')).strip()
            
            # Buscar registro correspondente na lista
            registro_encontrado = None
            indice_registro = -1
            
            for i, registro in enumerate(self.controle_mensal_dados):
                if str(registro.get('Numero_Instalacao', '')).strip() == numero_instalacao:
                    registro_encontrado = registro
                    indice_registro = i
                    break
            
            if not registro_encontrado:
                print(f"⚠️ Instalação {numero_instalacao} não encontrada na planilha relacionamento")
                return False
            
            # CONTROLE DE DUPLICATAS - Verificar se já foi processada
            status_atual = registro_encontrado.get('Status', 'Faltando')
            
            if status_atual == 'Recebida':
                print(f"🔄 DUPLICATA DETECTADA: Instalação {numero_instalacao} já processada este mês")
                print(f"   📅 Status atual: {status_atual}")
                print(f"   🚫 Ignorando fatura duplicada (padrão BRK)")
                
                # Registrar duplicata para relatório final
                self.estatisticas_duplicatas["faturas_duplicadas"] += 1
                self.estatisticas_duplicatas["instalacoes_reprocessadas"].append({
                    "instalacao": numero_instalacao,
                    "arquivo": dados_fatura.get('arquivo', 'N/A'),
                    "timestamp": datetime.now().isoformat()
                })
                
                return False  # Não processar duplicata
            else:
                print(f"🔄 ATUALIZAÇÃO: Instalação {numero_instalacao} - Status: {status_atual} → Recebida")
            
            # Atualizar dados no registro (FORMATO REAL ENEL - 18 colunas)
            
            # Campos básicos (conforme planilha real)
            valor_num = dados_fatura.get('valor_total_num')
            if valor_num is None:
                self.controle_mensal_dados[indice_registro]['Valor'] = 'ERRO_EXTRAÇÃO'
                print(f"❌ ERRO: Valor não extraído da fatura {dados_fatura.get('arquivo', 'N/A')}")
            else:
                # Formato brasileiro: R$ 126,37
                self.controle_mensal_dados[indice_registro]['Valor'] = f"R$ {valor_num:.2f}".replace('.', ',')
            
            # Competência: 06/2025
            self.controle_mensal_dados[indice_registro]['Competencia'] = dados_fatura.get('competencia', '')
            
            # Data_Emissao: 10/06/2025  
            self.controle_mensal_dados[indice_registro]['Data_Emissao'] = dados_fatura.get('data_emissao', '')
            
            # Nota_Fiscal: 718968230
            self.controle_mensal_dados[indice_registro]['Nota_Fiscal'] = dados_fatura.get('nota_fiscal', '')
            
            # Vencimento: 14/07/2025
            self.controle_mensal_dados[indice_registro]['Vencimento'] = dados_fatura.get('data_vencimento', '')
            
            # Consumo_kWh: 280,00 kWh (formato brasileiro)
            consumo = dados_fatura.get('consumo_kwh_num', 0)
            self.controle_mensal_dados[indice_registro]['Consumo_kWh'] = f"{consumo:.2f} kWh".replace('.', ',')
            
            # Sistema fotovoltaico (FORMATO REAL)
            self.controle_mensal_dados[indice_registro]['Sistema_Fotovoltaico'] = dados_fatura.get('sistema_fotovoltaico', 'Não')
            self.controle_mensal_dados[indice_registro]['Compensacao_TUSD'] = dados_fatura.get('compensacao_tusd_num', 0.0)
            self.controle_mensal_dados[indice_registro]['Compensacao_TE'] = dados_fatura.get('compensacao_te_num', 0.0)
            self.controle_mensal_dados[indice_registro]['Total_Compensacao'] = dados_fatura.get('total_compensacao', 0.0)
            
            # Calcular valor integral sem FV (valor + total compensação)
            if valor_num is not None:
                valor_integral = valor_num + dados_fatura.get('total_compensacao', 0.0)
                self.controle_mensal_dados[indice_registro]['Valor_Integral_Sem_FV'] = valor_integral
                
                # Calcular percentual economia FV
                if valor_integral > 0:
                    economia_percentual = (dados_fatura.get('total_compensacao', 0.0) / valor_integral) * 100
                    self.controle_mensal_dados[indice_registro]['Percentual_Economia_FV'] = round(economia_percentual, 2)
            
            # Calcular médias e diferenças percentuais (versão sem pandas)
            self._calcular_medias_consumo_sem_pandas(indice_registro, consumo)
            
            # Marcar como recebida
            self.controle_mensal_dados[indice_registro]['Status'] = 'Recebida'
            
            casa_oracao = self.controle_mensal_dados[indice_registro]['Casa de Oração']
            
            # Log do processamento
            if valor_num is not None:
                print(f"✅ Fatura processada: {casa_oracao} - R$ {valor_num:.2f}")
            else:
                print(f"❌ Fatura COM ERRO: {casa_oracao} - Valor não extraído")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro processando fatura incremental: {e}")
            return False
    
    def _calcular_medias_consumo_sem_pandas(self, indice_registro: int, consumo_atual: float):
        """
        Calcular médias e diferenças percentuais para uma instalação (versão sem pandas)
        
        Args:
            indice_registro: Índice do registro na lista
            consumo_atual: Consumo atual da fatura
        """
        try:
            # Por enquanto usar valores padrão (em produção viria do histórico do database)
            media_6_meses = consumo_atual  # Simplificado para validação
            diferenca_percentual = 0.0
            porcentagem_consumo = 100.0
            alerta_consumo = 'Normal'
            
            # Simular cálculo de diferença (em produção usar histórico real)
            if media_6_meses > 0:
                diferenca_percentual = ((consumo_atual - media_6_meses) / media_6_meses) * 100
                porcentagem_consumo = (consumo_atual / media_6_meses) * 100
                
                # USAR FUNÇÃO CENTRAL UNIFICADA - Uma única fonte de verdade
                from processor.classificador_consumo import determinar_tipo_alerta_consumo
                classificacao = determinar_tipo_alerta_consumo(consumo_atual, media_6_meses)
                alerta_consumo = classificacao['classificacao']
            
            # Atualizar registro
            self.controle_mensal_dados[indice_registro]['Media_6_Meses'] = round(media_6_meses, 2)
            self.controle_mensal_dados[indice_registro]['Diferenca_Percentual'] = round(diferenca_percentual, 2)
            self.controle_mensal_dados[indice_registro]['Porcentagem_Consumo'] = round(porcentagem_consumo, 2)
            self.controle_mensal_dados[indice_registro]['Alerta_Consumo'] = alerta_consumo
                
        except Exception as e:
            print(f"❌ Erro calculando médias: {e}")
            # Valores padrão em caso de erro
            self.controle_mensal_dados[indice_registro]['Media_6_Meses'] = 0.0
            self.controle_mensal_dados[indice_registro]['Diferenca_Percentual'] = 0.0
            self.controle_mensal_dados[indice_registro]['Porcentagem_Consumo'] = 100.0
            self.controle_mensal_dados[indice_registro]['Alerta_Consumo'] = 'Erro'
    
    def obter_status_controle(self) -> Dict:
        """
        Obter status atual do controle mensal
        
        Returns:
            Dict: Status detalhado
        """
        try:
            if self.planilha_controle_atual is None:
                return {"erro": "Planilha controle não carregada"}
            
            total_instalacoes = len(self.planilha_controle_atual)
            recebidas = len(self.planilha_controle_atual[self.planilha_controle_atual['Status'] == 'Recebida'])
            faltando = total_instalacoes - recebidas
            
            # Valor total processado
            valor_total = self.planilha_controle_atual[
                self.planilha_controle_atual['Status'] == 'Recebida'
            ]['Valor_Total'].sum()
            
            # Instalações faltantes
            faltantes = self.planilha_controle_atual[
                self.planilha_controle_atual['Status'] == 'Faltando'
            ][['Casa de Oração', 'Numero_Instalacao', 'Data_Vencimento']].to_dict('records')
            
            return {
                "competencia": f"{self.mes_atual:02d}/{self.ano_atual}",
                "total_instalacoes": total_instalacoes,
                "faturas_recebidas": recebidas,
                "faturas_faltando": faltando,
                "percentual_completo": round((recebidas / total_instalacoes) * 100, 1),
                "valor_total_processado": round(valor_total, 2),
                "instalacoes_faltantes": faltantes,
                "timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            return {"erro": str(e)}
    
    def _adicionar_aba_resumo_duplicatas(self, writer):
        """
        Adicionar aba de resumo com informações de duplicatas (igual BRK)
        
        Args:
            writer: ExcelWriter para adicionar a aba
        """
        try:
            # pandas removido - usa apenas openpyxl no Render
            # import pandas as pd
            from datetime import datetime
            
            # Obter estatísticas atuais
            status_controle = self.obter_status_controle()
            
            # Dados para a aba resumo
            resumo_data = [
                ["RESUMO DO PROCESSAMENTO ENEL", ""],
                ["", ""],
                ["Data/Hora do Processamento:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                ["Mês/Ano de Referência:", f"{self.mes_atual:02d}/{self.ano_atual}"],
                ["", ""],
                ["ESTATÍSTICAS GERAIS:", ""],
                ["Total de Instalações:", status_controle.get('total_instalacoes', 0)],
                ["Faturas Recebidas:", status_controle.get('faturas_recebidas', 0)],
                ["Faturas Faltando:", status_controle.get('faturas_faltando', 0)],
                ["Valor Total Processado:", f"R$ {status_controle.get('valor_total_processado', 0):.2f}"],
                ["Percentual Completo:", f"{status_controle.get('percentual_completo', 0):.1f}%"],
                ["", ""],
                ["CONTROLE DE DUPLICATAS (BRK Pattern):", ""],
                ["Emails Duplicados Ignorados:", self.estatisticas_duplicatas.get('emails_duplicados', 0)],
                ["Faturas Duplicadas Ignoradas:", self.estatisticas_duplicatas.get('faturas_duplicadas', 0)],
                ["", ""],
            ]
            
            # Adicionar detalhes de instalações reprocessadas se houver
            instalacoes_reprocessadas = self.estatisticas_duplicatas.get('instalacoes_reprocessadas', [])
            if instalacoes_reprocessadas:
                resumo_data.extend([
                    ["INSTALAÇÕES COM TENTATIVA DE REPROCESSAMENTO:", ""],
                    ["Instalação", "Arquivo PDF"]
                ])
                
                for item in instalacoes_reprocessadas[:10]:  # Máximo 10 para não poluir
                    resumo_data.append([item.get('instalacao', 'N/A'), item.get('arquivo', 'N/A')])
                    
                if len(instalacoes_reprocessadas) > 10:
                    resumo_data.append([f"... e mais {len(instalacoes_reprocessadas) - 10} registros", ""])
            
            # Adicionar rodapé
            resumo_data.extend([
                ["", ""],
                ["Sistema ENEL - Controle de Duplicatas Ativo", ""],
                ["Baseado no padrão BRK que funciona perfeitamente", ""]
            ])
            
            # Criar DataFrame do resumo
            df_resumo = pd.DataFrame(resumo_data, columns=['Campo', 'Valor'])
            
            # Adicionar aba
            df_resumo.to_excel(writer, sheet_name='Resumo Duplicatas', index=False)
            
            # Aplicar formatação à aba resumo
            workbook = writer.book
            worksheet_resumo = writer.sheets['Resumo Duplicatas']
            
            # Ajustar largura das colunas
            worksheet_resumo.column_dimensions['A'].width = 40
            worksheet_resumo.column_dimensions['B'].width = 30
            
            print(f"📄 Aba 'Resumo Duplicatas' adicionada (padrão BRK)")
            
        except Exception as e:
            print(f"⚠️ Erro adicionando aba resumo: {e}")
    
    def salvar_controle_mensal(self) -> bool:
        """Salvar planilha de controle atual no OneDrive"""
        try:
            if self.planilha_controle_atual is None:
                return False
            
            nome_arquivo = f"controle_enel_{self.ano_atual}_{self.mes_atual:02d}.xlsx"
            
            # Converter para bytes
            from io import BytesIO
            buffer = BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Aba principal - Controle ENEL
                self.planilha_controle_atual.to_excel(writer, sheet_name='Controle ENEL', index=False)
                
                # ABA DE RESUMO DE DUPLICATAS (igual BRK)
                self._adicionar_aba_resumo_duplicatas(writer)
                
                # Aplicar formatação básica
                workbook = writer.book
                worksheet = writer.sheets['Controle ENEL']
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            buffer.seek(0)
            arquivo_bytes = buffer.getvalue()
            
            # Upload para OneDrive na mesma pasta das faturas (IGUAL BRK)
            pasta_mes_id = self.onedrive.garantir_pasta_mes_ano(self.ano_atual, self.mes_atual)
            if not pasta_mes_id:
                print(f"❌ Erro: pasta do mês {self.mes_atual:02d}/{self.ano_atual} não encontrada")
                return False
                
            sucesso = self.onedrive.upload_arquivo(
                arquivo_bytes,
                nome_arquivo,
                pasta_mes_id  # Mesma pasta das faturas
            )
            
            if sucesso:
                print(f"💾 Controle mensal salvo: {nome_arquivo}")
                return True
            else:
                print(f"❌ Erro salvando controle mensal")
                return False
                
        except Exception as e:
            print(f"❌ Erro salvando controle: {e}")
            return False
    
    def _salvar_planilha_relacionamento(self) -> bool:
        """Salvar planilha de relacionamento no OneDrive"""
        try:
            if self.planilha_relacionamento is None:
                return False
            
            from io import BytesIO
            buffer = BytesIO()
            self.planilha_relacionamento.to_excel(buffer, index=False)
            buffer.seek(0)
            
            return self.onedrive.upload_arquivo(
                buffer.getvalue(),
                self.nome_planilha_relacionamento,
                self.onedrive.pasta_planilhas_id
            )
            
        except Exception as e:
            print(f"❌ Erro salvando planilha relacionamento: {e}")
            return False
    
    def carregar_planilha_relacionamento_sem_pandas(self) -> bool:
        """
        Carregar planilha de relacionamento ENEL SEM pandas (Render compatible)
        Baseado no padrão BRK que funciona perfeitamente
        
        Returns:
            bool: True se carregada com sucesso
        """
        try:
            if not EXCEL_AVAILABLE:
                print("⚠️ openpyxl não disponível")
                return False
            
            if not self.onedrive.pasta_enel_id:
                print("❌ Pasta raiz ENEL OneDrive não configurada")
                return False
            
            print(f"📁 Carregando planilha {self.nome_planilha_relacionamento} da raiz ENEL (igual BRK)...")
            
            # Buscar arquivo na pasta raiz ENEL (igual BRK)
            planilha_bytes = self._baixar_arquivo_onedrive(
                self.onedrive.pasta_enel_id, 
                self.nome_planilha_relacionamento
            )
            
            if not planilha_bytes:
                print(f"⚠️ Planilha {self.nome_planilha_relacionamento} não encontrada")
                return self._criar_planilha_relacionamento_exemplo_sem_pandas()
            
            # Processar Excel manualmente (sem pandas)
            registros = self._processar_excel_relacionamento_manual(planilha_bytes)
            
            if not registros:
                print(f"❌ Nenhum registro encontrado na planilha")
                return False
            
            # Converter para estrutura interna (sem pandas)
            self.relacionamentos_dados = registros
            
            print(f"✅ PLANILHA DE RELACIONAMENTO CARREGADA SEM pandas!")
            print(f"   📄 Registros: {len(registros)}")
            print(f"   🏠 Casas: {len(set(r.get('Casa', '') for r in registros))}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro carregando planilha sem pandas: {e}")
            return False
    
    def _processar_excel_relacionamento_manual(self, excel_bytes) -> List[Dict]:
        """
        Processar Excel de relacionamento manualmente (sem pandas)
        Estrutura: A=Casa, B=Instalação, C=Vencimento
        
        Args:
            excel_bytes: Bytes do arquivo Excel
            
        Returns:
            List[Dict]: Registros processados
        """
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            import io
            
            registros = []
            
            # Excel (.xlsx) é um arquivo ZIP com XMLs internos
            with zipfile.ZipFile(io.BytesIO(excel_bytes), 'r') as zip_file:
                
                # 1. Ler shared strings (textos da planilha)
                shared_strings = []
                try:
                    with zip_file.open('xl/sharedStrings.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        
                        # Namespace do Excel
                        ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                        for si in root.findall('.//ss:si', ns):
                            t_elem = si.find('ss:t', ns)
                            if t_elem is not None:
                                shared_strings.append(t_elem.text or '')
                            else:
                                shared_strings.append('')
                except:
                    pass  # Planilha pode não ter shared strings
                
                # 2. Ler dados da primeira worksheet
                try:
                    with zip_file.open('xl/worksheets/sheet1.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        
                        ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                        # Processar linhas (pular cabeçalho)
                        rows = root.findall('.//ss:row', ns)
                        for row_idx, row in enumerate(rows[1:], 2):  # Pular linha 1 (cabeçalho)
                            cells = row.findall('ss:c', ns)
                            
                            casa = ''
                            instalacao = ''
                            vencimento = ''
                            
                            for cell in cells:
                                cell_ref = cell.get('r', '')
                                col_letter = ''.join([c for c in cell_ref if c.isalpha()])
                                
                                # Obter valor da célula
                                valor = ''
                                v_elem = cell.find('ss:v', ns)
                                if v_elem is not None:
                                    if cell.get('t') == 's':  # String compartilhada
                                        try:
                                            idx = int(v_elem.text)
                                            if 0 <= idx < len(shared_strings):
                                                valor = shared_strings[idx]
                                        except (ValueError, IndexError):
                                            valor = v_elem.text or ''
                                    else:
                                        valor = v_elem.text or ''
                                
                                # Mapear colunas
                                if col_letter == 'A':
                                    casa = str(valor).strip()
                                elif col_letter == 'B':
                                    instalacao = str(valor).strip()
                                elif col_letter == 'C':
                                    vencimento = str(valor).strip()
                            
                            # Adicionar registro se válido
                            if casa and instalacao:
                                registros.append({
                                    'Casa': casa,
                                    'Instalacao': instalacao,
                                    'Vencimento': vencimento or '0'
                                })
                
                except Exception as e:
                    print(f"⚠️ Erro processando worksheet: {e}")
            
            return registros
            
        except Exception as e:
            print(f"❌ Erro processando Excel manual: {e}")
            return []
    
    def _criar_planilha_relacionamento_exemplo_sem_pandas(self) -> bool:
        """
        Criar planilha de relacionamento exemplo SEM pandas
        """
        try:
            if not EXCEL_AVAILABLE:
                return False
            
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relacionamento ENEL"
            
            # Cabeçalhos
            ws.cell(row=1, column=1, value="Casa de Oração")
            ws.cell(row=1, column=2, value="Número Instalação")
            ws.cell(row=1, column=3, value="Dia Vencimento")
            
            # Dados de exemplo
            exemplos = [
                ("Casa Exemplo 1", "12345678", "10"),
                ("Casa Exemplo 2", "87654321", "15"),
                ("Casa Exemplo 3", "11223344", "20")
            ]
            
            for row_idx, (casa, instalacao, vencimento) in enumerate(exemplos, 2):
                ws.cell(row=row_idx, column=1, value=casa)
                ws.cell(row=row_idx, column=2, value=instalacao)
                ws.cell(row=row_idx, column=3, value=vencimento)
            
            # Salvar no OneDrive
            buffer = io.BytesIO()
            wb.save(buffer)
            planilha_bytes = buffer.getvalue()
            buffer.close()
            
            sucesso = self._upload_arquivo_onedrive(
                self.onedrive.pasta_enel_id,  # Pasta raiz ENEL (igual BRK)
                self.nome_planilha_relacionamento,
                planilha_bytes
            )
            
            if sucesso:
                # Definir dados exemplo na memória
                self.relacionamentos_dados = [
                    {'Casa': casa, 'Instalacao': instalacao, 'Vencimento': vencimento}
                    for casa, instalacao, vencimento in exemplos
                ]
                print(f"✅ Planilha exemplo criada SEM pandas!")
                return True
            
            return False
            
        except Exception as e:
            print(f"❌ Erro criando exemplo sem pandas: {e}")
            return False
    
    def _carregar_planilha_relacionamento_com_pandas(self) -> bool:
        """
        DEPRECATED: Método removido - usar apenas OneDrive
        """
        if not PANDAS_AVAILABLE:
            return False
            
        try:
            if not self.onedrive.pasta_enel_id:
                print("❌ Pasta raiz ENEL OneDrive não configurada")
                return False
            
            # Buscar arquivo na pasta raiz ENEL (igual BRK)
            planilha_bytes = self._baixar_arquivo_onedrive(
                self.onedrive.pasta_enel_id, 
                self.nome_planilha_relacionamento
            )
            
            if not planilha_bytes:
                print(f"⚠️ Planilha {self.nome_planilha_relacionamento} não encontrada")
                return self._criar_planilha_relacionamento_exemplo_com_pandas()
            
            # Carregar em pandas
            self.planilha_relacionamento = pd.read_excel(planilha_bytes)
            
            print(f"✅ Planilha relacionamento carregada COM pandas: {len(self.planilha_relacionamento)} registros")
            return True
            
        except Exception as e:
            print(f"❌ Erro carregando com pandas: {e}")
            return False
    
    def _baixar_arquivo_onedrive(self, pasta_id: str, nome_arquivo: str) -> Optional[bytes]:
        """Baixar arquivo do OneDrive"""
        try:
            if not self.auth.access_token:
                return None
            
            headers = self.auth.obter_headers_autenticados()
            
            # Buscar arquivo na pasta
            url = f"https://graph.microsoft.com/v1.0/me/drive/items/{pasta_id}/children"
            params = {"$filter": f"name eq '{nome_arquivo}'"}
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                items = response.json().get('value', [])
                
                if items:
                    file_id = items[0]['id']
                    
                    # Baixar conteúdo
                    content_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/content"
                    content_response = requests.get(content_url, headers=headers, timeout=60)
                    
                    if content_response.status_code == 200:
                        return content_response.content
            
            return None
            
        except Exception as e:
            print(f"❌ Erro baixando {nome_arquivo}: {e}")
            return None
    
    def _calcular_medias_consumo(self, idx):
        """
        Calcular médias e diferenças percentuais para uma instalação
        Usa histórico de consumo dos meses anteriores
        
        Args:
            idx: Índice da linha na planilha controle
        """
        try:
            if not PANDAS_AVAILABLE:
                # Sem pandas, usar valores padrão
                self.planilha_controle_atual.loc[idx, 'Media_6_Meses'] = 0.0
                self.planilha_controle_atual.loc[idx, 'Diferenca_Percentual'] = 0.0
                self.planilha_controle_atual.loc[idx, 'Porcentagem_Consumo'] = 100.0
                self.planilha_controle_atual.loc[idx, 'Alerta_Consumo'] = 'N/A'
                return
            
            # Obter dados da instalação
            numero_instalacao = self.planilha_controle_atual.loc[idx, 'Numero_Instalacao']
            consumo_atual = self.planilha_controle_atual.loc[idx, 'Consumo_kWh']
            
            if not numero_instalacao or consumo_atual == 0:
                return
            
            # Simular histórico (em produção viria do OneDrive)
            # Por enquanto usar consumo atual como base
            historico_consumo = [
                {"mes_ano": f"{self.mes_atual:02d}/{self.ano_atual}", "consumo": consumo_atual, "tipo": "ATUAL"}
            ]
            
            # Calcular usando função centralizada
            consumo_calc, media_6_meses, diferenca_percentual = calcular_media_e_diferenca_enel(historico_consumo)
            
            # Atualizar planilha
            self.planilha_controle_atual.loc[idx, 'Media_6_Meses'] = round(media_6_meses, 2)
            self.planilha_controle_atual.loc[idx, 'Diferenca_Percentual'] = round(diferenca_percentual, 2)
            
            # Calcular porcentagem de consumo
            if media_6_meses > 0:
                porcentagem_consumo = (consumo_atual / media_6_meses) * 100
                self.planilha_controle_atual.loc[idx, 'Porcentagem_Consumo'] = round(porcentagem_consumo, 2)
                
                # USAR FUNÇÃO CENTRAL UNIFICADA - Uma única fonte de verdade
                from processor.classificador_consumo import determinar_tipo_alerta_consumo
                classificacao = determinar_tipo_alerta_consumo(consumo_atual, media_6_meses)
                self.planilha_controle_atual.loc[idx, 'Alerta_Consumo'] = classificacao['classificacao']
            else:
                self.planilha_controle_atual.loc[idx, 'Porcentagem_Consumo'] = 100.0
                self.planilha_controle_atual.loc[idx, 'Alerta_Consumo'] = 'Sem histórico'
                
        except Exception as e:
            print(f"❌ Erro calculando médias para instalação {numero_instalacao}: {e}")
            # Valores padrão em caso de erro
            self.planilha_controle_atual.loc[idx, 'Media_6_Meses'] = 0.0
            self.planilha_controle_atual.loc[idx, 'Diferenca_Percentual'] = 0.0
            self.planilha_controle_atual.loc[idx, 'Porcentagem_Consumo'] = 100.0
            self.planilha_controle_atual.loc[idx, 'Alerta_Consumo'] = 'Erro'
    
    # ============================================================================
    # INTEGRAÇÃO COM DATABASE
    # ============================================================================
    
    def salvar_fatura_no_database(self, dados_fatura: Dict, email_id: str, pdf_content: bytes = None) -> bool:
        """
        Salvar fatura no database SQLite
        
        Args:
            dados_fatura: Dados extraídos da fatura
            email_id: ID do email original
            pdf_content: Conteúdo do PDF
            
        Returns:
            bool: True se salvo com sucesso
        """
        try:
            if not self.database:
                print("⚠️ Database não disponível")
                return False
            
            return self.database.inserir_fatura(dados_fatura, email_id, pdf_content)
            
        except Exception as e:
            print(f"❌ Erro salvando fatura no database: {e}")
            return False
    
    def obter_dados_database_para_planilha(self, numero_instalacao: str = None) -> List[Dict]:
        """
        Obter dados do database para gerar planilhas
        
        Args:
            numero_instalacao: Filtrar por instalação específica (opcional)
            
        Returns:
            List[Dict]: Dados formatados para planilha
        """
        try:
            if not self.database:
                print("⚠️ Database não disponível")
                return []
            
            if numero_instalacao:
                return self.database.buscar_por_instalacao(numero_instalacao)
            else:
                # Buscar todas as faturas do database
                estatisticas = self.database.obter_estatisticas()
                print(f"📊 Database tem {estatisticas.get('total_faturas', 0)} faturas")
                return []
                
        except Exception as e:
            print(f"❌ Erro obtendo dados do database: {e}")
            return []
    
    def gerar_planilha_excel_sem_pandas(self) -> bool:
        """
        Gerar planilha Excel SEM pandas usando apenas openpyxl
        Formato exato da planilha real ENEL (18 colunas)
        
        Returns:
            bool: True se gerada com sucesso
        """
        try:
            if not EXCEL_AVAILABLE:
                print("❌ openpyxl não disponível")
                return False
                
            if not self.controle_mensal_dados:
                print("❌ Dados do controle mensal não disponíveis")
                return False
            
            print(f"📊 Gerando planilha Excel SEM pandas...")
            
            # Criar workbook
            wb = openpyxl.Workbook()
            
            # Remover sheet padrão
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Criar aba principal
            ws_principal = wb.create_sheet("Faturas ENEL", 0)
            
            # 18 colunas EXATAS conforme planilha real
            cabecalhos = [
                'Casa de Oração', 'Competencia', 'Data_Emissao', 'Nota_Fiscal',
                'Vencimento', 'Valor', 'Consumo_kWh', 'Media_6_Meses', 
                'Diferenca_Percentual', 'Porcentagem_Consumo', 'Alerta_Consumo',
                'Sistema_Fotovoltaico', 'Compensacao_TUSD', 'Compensacao_TE',
                'Total_Compensacao', 'Valor_Integral_Sem_FV', 'Percentual_Economia_FV',
                'Numero_Instalacao'
            ]
            
            # Escrever cabeçalhos
            for col_idx, header in enumerate(cabecalhos, 1):
                cell = ws_principal.cell(row=1, column=col_idx, value=header)
                # Formatação do cabeçalho
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
            # Separar dados por grupos (PIA vs outros)
            dados_pia = []
            dados_outros = []
            
            for registro in self.controle_mensal_dados:
                casa = registro.get('Casa de Oração', '').upper()
                if 'PIA' in casa:
                    dados_pia.append(registro)
                else:
                    dados_outros.append(registro)
            
            linha_atual = 2
            
            # Função para escrever dados de um grupo
            def escrever_grupo(dados_grupo, nome_grupo, linha_inicio):
                linha_atual = linha_inicio
                
                # Cabeçalho do grupo
                if dados_grupo:
                    ws_principal.cell(row=linha_atual, column=1, value=f"=== GRUPO {nome_grupo} ===")
                    ws_principal.cell(row=linha_atual, column=1).font = Font(bold=True, color="0066CC")
                    linha_atual += 1
                
                # Dados do grupo
                subtotal_valor = 0.0
                for registro in dados_grupo:
                    for col_idx, header in enumerate(cabecalhos, 1):
                        valor = registro.get(header, '')
                        
                        # Formatação especial para valores monetários
                        if header == 'Valor' and isinstance(valor, str) and valor.startswith('R$'):
                            try:
                                # Extrair valor numérico para subtotal
                                valor_num = float(valor.replace('R$ ', '').replace(',', '.'))
                                subtotal_valor += valor_num
                            except:
                                pass
                        
                        ws_principal.cell(row=linha_atual, column=col_idx, value=valor)
                    
                    linha_atual += 1
                
                # Subtotal do grupo
                if dados_grupo and subtotal_valor > 0:
                    ws_principal.cell(row=linha_atual, column=1, value=f"SUBTOTAL {nome_grupo}")
                    ws_principal.cell(row=linha_atual, column=1).font = Font(bold=True)
                    ws_principal.cell(row=linha_atual, column=6, value=f"R$ {subtotal_valor:.2f}".replace('.', ','))
                    ws_principal.cell(row=linha_atual, column=6).font = Font(bold=True)
                    linha_atual += 2  # Espaço após subtotal
                
                return linha_atual
            
            # Escrever grupo PIA
            linha_atual = escrever_grupo(dados_pia, "PIA", linha_atual)
            
            # Escrever outros grupos
            linha_atual = escrever_grupo(dados_outros, "OUTROS", linha_atual)
            
            # Total geral
            total_geral = sum([
                float(reg.get('Valor', '').replace('R$ ', '').replace(',', '.')) 
                for reg in self.controle_mensal_dados 
                if reg.get('Valor', '').startswith('R$')
            ])
            
            if total_geral > 0:
                ws_principal.cell(row=linha_atual, column=1, value="TOTAL GERAL")
                ws_principal.cell(row=linha_atual, column=1).font = Font(bold=True, color="FF0000")
                ws_principal.cell(row=linha_atual, column=6, value=f"R$ {total_geral:.2f}".replace('.', ','))
                ws_principal.cell(row=linha_atual, column=6).font = Font(bold=True, color="FF0000")
            
            # Ajustar larguras das colunas
            for col_idx in range(1, len(cabecalhos) + 1):
                col_letter = get_column_letter(col_idx)
                # Largura baseada no conteúdo
                if col_idx == 1:  # Casa de Oração - mais larga
                    ws_principal.column_dimensions[col_letter].width = 25
                elif col_idx in [6, 7]:  # Valor e Consumo - largura média
                    ws_principal.column_dimensions[col_letter].width = 15
                elif col_idx == 18:  # Número Instalação
                    ws_principal.column_dimensions[col_letter].width = 12
                else:
                    ws_principal.column_dimensions[col_letter].width = 12
            
            # Adicionar aba de resumo
            self._adicionar_aba_resumo_sem_pandas(wb)
            
            # Salvar arquivo
            nome_arquivo = f"Faturas_ENEL_Completo_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            # Converter para bytes
            import io
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            arquivo_bytes = buffer.getvalue()
            
            # Upload para OneDrive na mesma pasta das faturas (IGUAL BRK)
            pasta_mes_id = self.onedrive.garantir_pasta_mes_ano(self.ano_atual, self.mes_atual)
            if pasta_mes_id:
                sucesso = self._upload_arquivo_onedrive(
                    pasta_mes_id,  # Mesma pasta das faturas
                    nome_arquivo,
                    arquivo_bytes
                )
                
                if sucesso:
                    print(f"✅ Planilha Excel gerada e salva no OneDrive: {nome_arquivo}")
                    print(f"   📄 Registros processados: {len(self.controle_mensal_dados)}")
                    print(f"   💰 Total geral: R$ {total_geral:.2f}")
                    return True
                else:
                    print(f"❌ Erro salvando planilha no OneDrive")
                    return False
            else:
                print(f"⚠️ OneDrive não disponível - planilha criada apenas localmente")
                return True
            
        except Exception as e:
            print(f"❌ Erro gerando planilha Excel: {e}")
            return False
    
    def _adicionar_aba_resumo_sem_pandas(self, workbook):
        """
        Adicionar aba de resumo SEM pandas (apenas openpyxl)
        
        Args:
            workbook: Workbook do openpyxl para adicionar a aba
        """
        try:
            ws_resumo = workbook.create_sheet("Resumo", 1)
            
            # Dados do resumo
            linha = 1
            resumo_dados = [
                ("RESUMO DO PROCESSAMENTO ENEL", ""),
                ("", ""),
                ("Data/Hora do Processamento:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                ("Mês/Ano de Referência:", f"{self.mes_atual:02d}/{self.ano_atual}"),
                ("", ""),
                ("ESTATÍSTICAS GERAIS:", ""),
                ("Total de Instalações:", len(self.controle_mensal_dados)),
                ("Faturas Recebidas:", len([r for r in self.controle_mensal_dados if r.get('Status') == 'Recebida'])),
                ("Faturas Faltando:", len([r for r in self.controle_mensal_dados if r.get('Status') == 'Faltando'])),
                ("", ""),
                ("CONTROLE DE DUPLICATAS (BRK Pattern):", ""),
                ("Emails Duplicados Ignorados:", self.estatisticas_duplicatas.get('emails_duplicados', 0)),
                ("Faturas Duplicadas Ignoradas:", self.estatisticas_duplicatas.get('faturas_duplicadas', 0)),
                ("", ""),
                ("Sistema ENEL - Controle de Duplicatas Ativo", ""),
                ("Baseado no padrão BRK que funciona perfeitamente", "")
            ]
            
            for campo, valor in resumo_dados:
                ws_resumo.cell(row=linha, column=1, value=campo)
                ws_resumo.cell(row=linha, column=2, value=valor)
                
                # Formatação especial para títulos
                if "RESUMO" in str(campo) or "ESTATÍSTICAS" in str(campo) or "CONTROLE" in str(campo):
                    ws_resumo.cell(row=linha, column=1).font = Font(bold=True, color="0066CC")
                
                linha += 1
            
            # Ajustar larguras
            ws_resumo.column_dimensions['A'].width = 40
            ws_resumo.column_dimensions['B'].width = 30
            
            print(f"📄 Aba 'Resumo' adicionada SEM pandas")
            
        except Exception as e:
            print(f"⚠️ Erro adicionando aba resumo: {e}")
    
    def _upload_arquivo_onedrive(self, pasta_id: str, nome_arquivo: str, conteudo_bytes: bytes) -> bool:
        """
        Upload de arquivo para OneDrive
        
        Args:
            pasta_id: ID da pasta no OneDrive
            nome_arquivo: Nome do arquivo
            conteudo_bytes: Conteúdo em bytes
            
        Returns:
            bool: True se upload realizado com sucesso
        """
        try:
            if not self.auth.access_token:
                return False
            
            headers = self.auth.obter_headers_autenticados()
            headers['Content-Type'] = 'application/octet-stream'
            
            # URL para upload
            url = f"https://graph.microsoft.com/v1.0/me/drive/items/{pasta_id}:/{nome_arquivo}:/content"
            
            response = requests.put(url, headers=headers, data=conteudo_bytes, timeout=120)
            
            if response.status_code in [200, 201]:
                return True
            else:
                print(f"❌ Erro upload OneDrive: {response.status_code}")
                return False
            
        except Exception as e:
            print(f"❌ Erro fazendo upload: {e}")
            return False