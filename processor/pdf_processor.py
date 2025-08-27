#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📁 ARQUIVO: processor/pdf_processor.py
💾 FUNÇÃO: Processador de PDFs ENEL
🔧 DESCRIÇÃO: Remoção de proteção, extração de dados, geração de planilhas
👨‍💼 AUTOR: Adaptado do sistema desktop ENEL
🔒 BASEADO EM: funcionalidade do processor.py original
"""

import os
import json
from datetime import datetime
from typing import Dict, List, Optional
import re
from pathlib import Path

# Imports condicionais para processamento de PDF
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️ PyPDF2 não disponível - funcionalidade de PDF limitada")

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    print("⚠️ pdfplumber não disponível - extração de dados limitada")

# Remover dependência do pandas - funciona sem pandas no Render
EXCEL_AVAILABLE = True  # Sempre True pois processaremos manualmente

# Imports opcionais - pandas removido para Render
try:
    # import pandas as pd  # Removido - Render não suporta
    import openpyxl
    PANDAS_AVAILABLE = False  # Forçado False - sem pandas
except ImportError:
    PANDAS_AVAILABLE = False

class PDFProcessor:
    """
    Processador de PDFs ENEL (VERSÃO SEM PANDAS - RENDER COMPATÍVEL)
    
    Responsabilidades:
    - Remover proteção de PDFs (senha padrão ENEL)
    - Extrair dados das faturas SEM pandas
    - Processar dados manualmente
    
    BASEADO NO PADRÃO BRK que funciona perfeitamente no Render
    """
    
    def __init__(self):
        """Inicializar processador de PDFs ENEL"""
        self.senha_padrao = os.getenv('SENHA_PDF_ENEL', '05150')
        self.storage_dir = "/opt/render/project/storage"
        self.pdfs_dir = os.path.join(self.storage_dir, "pdfs_enel")
        self.planilhas_dir = os.path.join(self.storage_dir, "planilhas_enel")
        
        # Criar diretórios
        os.makedirs(self.pdfs_dir, exist_ok=True)
        os.makedirs(self.planilhas_dir, exist_ok=True)
        
        # Cache de dados extraídos
        self.dados_extraidos = []
        
        print(f"📄 PDFProcessor ENEL inicializado")
        print(f"   Senha padrão: configurada")
        print(f"   PyPDF2: {'✅ disponível' if PDF_AVAILABLE else '❌ indisponível'}")
        print(f"   pdfplumber: {'✅ disponível' if PDFPLUMBER_AVAILABLE else '❌ indisponível'}")
        print(f"   Excel: {'✅ disponível' if EXCEL_AVAILABLE else '❌ indisponível'}")
    
    def remover_protecao_pdf(self, caminho_pdf: str) -> bool:
        """
        Remover proteção de PDF com senha padrão ENEL
        
        Args:
            caminho_pdf (str): Caminho do PDF protegido
            
        Returns:
            bool: True se removido com sucesso
        """
        if not PDF_AVAILABLE:
            print("⚠️ PyPDF2 não disponível para remoção de proteção")
            return False
        
        try:
            if not os.path.exists(caminho_pdf):
                print(f"❌ Arquivo não encontrado: {caminho_pdf}")
                return False
            
            # Tentar abrir PDF
            with open(caminho_pdf, 'rb') as arquivo:
                reader = PyPDF2.PdfReader(arquivo)
                
                # Verificar se está protegido
                if not reader.is_encrypted:
                    print(f"📄 PDF já desprotegido: {os.path.basename(caminho_pdf)}")
                    return True
                
                # Tentar desproteger com senha padrão
                if reader.decrypt(self.senha_padrao):
                    # Criar novo PDF sem proteção
                    writer = PyPDF2.PdfWriter()
                    
                    for page_num in range(len(reader.pages)):
                        page = reader.pages[page_num]
                        writer.add_page(page)
                    
                    # Salvar versão desprotegida (sobrescrever original)
                    caminho_temp = caminho_pdf + '.temp'
                    with open(caminho_temp, 'wb') as output_file:
                        writer.write(output_file)
                    
                    # Substituir arquivo original
                    os.replace(caminho_temp, caminho_pdf)
                    
                    print(f"🔓 PDF desprotegido: {os.path.basename(caminho_pdf)}")
                    return True
                else:
                    print(f"❌ Senha incorreta para: {os.path.basename(caminho_pdf)}")
                    return False
                    
        except Exception as e:
            print(f"❌ Erro removendo proteção de {os.path.basename(caminho_pdf)}: {e}")
            return False
    
    def extrair_dados_fatura(self, caminho_pdf: str) -> Optional[Dict]:
        """
        Extrair dados de uma fatura ENEL
        
        Args:
            caminho_pdf (str): Caminho do PDF da fatura
            
        Returns:
            Dict: Dados extraídos ou None se falhar
        """
        if not PDFPLUMBER_AVAILABLE:
            print("⚠️ pdfplumber não disponível para extração de dados")
            return None
        
        try:
            if not os.path.exists(caminho_pdf):
                return None
            
            dados_fatura = {
                "arquivo": os.path.basename(caminho_pdf),
                "caminho": caminho_pdf,
                "data_processamento": datetime.now().isoformat()
            }
            
            with pdfplumber.open(caminho_pdf) as pdf:
                texto_completo = ""
                
                # Extrair texto de todas as páginas
                for pagina in pdf.pages:
                    texto_pagina = pagina.extract_text()
                    if texto_pagina:
                        texto_completo += texto_pagina + "\n"
                
                if not texto_completo.strip():
                    print(f"⚠️ Nenhum texto extraído de: {os.path.basename(caminho_pdf)}")
                    return dados_fatura
                
                # Extrair dados específicos ENEL usando regex
                dados_fatura.update(self._extrair_campos_enel(texto_completo))
                
                print(f"📊 Dados extraídos de: {os.path.basename(caminho_pdf)}")
                return dados_fatura
                
        except Exception as e:
            print(f"❌ Erro extraindo dados de {os.path.basename(caminho_pdf)}: {e}")
            return None
    
    def _extrair_campos_enel(self, texto: str) -> Dict:
        """
        Extrair campos específicos das faturas ENEL
        
        Args:
            texto (str): Texto extraído do PDF
            
        Returns:
            Dict: Campos extraídos
        """
        campos = {}
        
        # Padrões regex para dados ENEL (adaptados do sistema original)
        padroes = {
            'numero_instalacao': [
                r'Instalação\s*:?\s*(\d+)',
                r'Nº\s*Instalação\s*:?\s*(\d+)',
                r'Instalacao\s*:?\s*(\d+)'
            ],
            'nota_fiscal': [
                r'Nota\s*Fiscal\s*:?\s*(\d+)',
                r'NF\s*:?\s*(\d+)',
                r'N\.?\s*F\.?\s*:?\s*(\d+)'
            ],
            'data_emissao': [
                r'Data\s*de\s*Emissão\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})',
                r'Emissão\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})',
                r'Data\s*Emissão\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})',
                r'Emitida\s*em\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})',
                r'Fatura\s*emitida\s*em\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})'
            ],
            'data_vencimento': [
                r'Vencimento\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})',
                r'Vence\s*em\s*:?\s*(\d{2}[\/\-]\d{2}[\/\-]\d{4})'
            ],
            'valor_total': [
                r'Valor\s*Total\s*:?\s*R\$\s*([\d.,]+)',
                r'Total\s*a\s*Pagar\s*:?\s*R\$\s*([\d.,]+)',
                r'TOTAL\s*:?\s*R\$\s*([\d.,]+)',
                r'VALOR\s*DO\s*DOCUMENTO\s*:?\s*R\$\s*([\d.,]+)',
                r'TOTAL\s+([\d.,]+)\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+',
                r'R\$\s*([\d.,]+)'
            ],
            'consumo_kwh': [
                r'Consumo\s*:?\s*(\d+)\s*kWh',
                r'Energia\s*Ativa\s*:?\s*(\d+)\s*kWh'
            ],
            'competencia': [
                r'Competência\s*:?\s*(\d{2}[\/\-]\d{4})',
                r'Referente\s*a\s*:?\s*(\d{2}[\/\-]\d{4})'
            ],
            'compensacao_tusd': [
                # Padrões REAIS da fatura ENEL com fotovoltaico
                r'EN\s*COMP\s*ATV\s*TU.*?(\d+,\d+)\s*-?([\d.,]+)',  # EN COMP ATV TU ... -75,60
                r'COMP.*?TU.*?-?([\d.,]+)',                        # COMP ... TU ... -75,60
                r'Compensação\s*TUSD\s*:?\s*-?\s*R?\$?\s*([\d.,]+)', # Formato tradicional
                r'TUSD.*?-?([\d.,]+)',                             # TUSD ... -75,60
            ],
            'compensacao_te': [
                # Padrões REAIS da fatura ENEL com fotovoltaico  
                r'EN\s*COMP\s*ATV\s*TE.*?(\d+,\d+)\s*-?([\d.,]+)',  # EN COMP ATV TE ... -82,51
                r'COMP.*?TE.*?-?([\d.,]+)',                        # COMP ... TE ... -82,51
                r'Compensação\s*TE\s*:?\s*-?\s*R?\$?\s*([\d.,]+)', # Formato tradicional
                r'TE.*?-?([\d.,]+)',                               # TE ... -82,51
            ],
            'energia_injetada': [
                r'Energia\s*Injetada\s*:?\s*(\d+)\s*kWh',
                r'Injetada\s*:?\s*(\d+)\s*kWh',
                r'Geração\s*:?\s*(\d+)\s*kWh',
                r'Energia\s*Gerada\s*:?\s*(\d+)\s*kWh'
            ],
            'energia_compensada': [
                # Padrão REAL: EN COMP ATV TU OUC05/25MPTGD2 KWH 239,182
                r'EN\s*COMP\s*ATV.*?KWH\s*([\d.,]+)',              # EN COMP ATV ... KWH 239,182
                r'Energia\s*Compensada\s*:?\s*([\d.,]+)\s*kWh',    # Formato tradicional
                r'Compensada\s*:?\s*([\d.,]+)\s*kWh',              # Compensada: 239,182 kWh
                r'COMP.*?(\d+,\d+)\s*kWh',                         # COMP ... 239,182 kWh
            ],
            'saldo_creditos': [
                # Saldo atualizado de energia em kWh: Ativa: 1.450,0
                r'Saldo\s*atualizado.*?Ativa\s*:\s*([\d.,]+)',     # Saldo ... Ativa: 1.450,0
                r'Saldo.*?energia.*?([\d.,]+)',                    # Saldo energia ... 1.450,0
            ]
        }
        
        # Aplicar padrões
        for campo, padroes_campo in padroes.items():
            for padrao in padroes_campo:
                match = re.search(padrao, texto, re.IGNORECASE | re.MULTILINE)
                if match:
                    valor = match.group(1).strip()
                    campos[campo] = valor
                    break
            
            # Valor padrão se não encontrado
            if campo not in campos:
                campos[campo] = "N/A"
        
        # Limpeza e conversões específicas - BASEADO NA SOLUÇÃO BRK
        # Converter valor numérico com tratamento seguro de formato brasileiro
        if campos.get('valor_total', 'N/A') != 'N/A':
            try:
                valor_str = campos['valor_total'].strip()
                
                # TRATAMENTO ROBUSTO (baseado na experiência BRK):
                # Verificar se tem vírgula (formato brasileiro padrão)
                if ',' in valor_str:
                    # Formato brasileiro: "1.126,37" ou "126,37"
                    # Remove pontos (separadores de milhares) e troca vírgula por ponto
                    valor_limpo = valor_str.replace('.', '').replace(',', '.')
                elif '.' in valor_str and len(valor_str.split('.')[-1]) == 2:
                    # Formato americano com centavos: "126.37" 
                    # Manter como está
                    valor_limpo = valor_str
                elif '.' in valor_str:
                    # Formato com separador de milhares sem centavos: "1.126"
                    # Remover pontos
                    valor_limpo = valor_str.replace('.', '')
                else:
                    # Sem separadores: "126"
                    valor_limpo = valor_str
                
                campos['valor_total_num'] = float(valor_limpo)
                print(f"💰 Valor convertido: '{campos['valor_total']}' → {campos['valor_total_num']}")
                
            except Exception as e:
                print(f"❌ ERRO CRÍTICO: Falha convertendo valor '{campos.get('valor_total')}': {e}")
                campos['valor_total_num'] = None
                campos['erro_extracao_valor'] = f"Erro conversão: {campos.get('valor_total')}"
        else:
            print(f"❌ ERRO CRÍTICO: Valor total não encontrado na fatura")
            campos['valor_total_num'] = None
            campos['erro_extracao_valor'] = "Valor total não encontrado"
        
        if campos.get('consumo_kwh', 'N/A') != 'N/A':
            try:
                campos['consumo_kwh_num'] = int(campos['consumo_kwh'])
            except:
                campos['consumo_kwh_num'] = 0
        else:
            campos['consumo_kwh_num'] = 0
        
        # Conversões para campos de energia fotovoltaica - TRATAMENTO SEGURO BRK
        def converter_valor_monetario(valor_str):
            """Conversão segura baseada na solução BRK"""
            if not valor_str or valor_str == 'N/A':
                return 0.0
            try:
                valor_str = valor_str.strip()
                if ',' in valor_str:
                    return float(valor_str.replace('.', '').replace(',', '.'))
                elif '.' in valor_str and len(valor_str.split('.')[-1]) == 2:
                    return float(valor_str)
                elif '.' in valor_str:
                    return float(valor_str.replace('.', ''))
                else:
                    return float(valor_str)
            except:
                return 0.0
        
        # Compensação TUSD
        campos['compensacao_tusd_num'] = converter_valor_monetario(campos.get('compensacao_tusd'))
        
        # Compensação TE  
        campos['compensacao_te_num'] = converter_valor_monetario(campos.get('compensacao_te'))
        
        # Energia injetada
        if campos.get('energia_injetada', 'N/A') != 'N/A':
            try:
                campos['energia_injetada_num'] = int(campos['energia_injetada'])
            except:
                campos['energia_injetada_num'] = 0
        else:
            campos['energia_injetada_num'] = 0
            
        # Energia compensada
        if campos.get('energia_compensada', 'N/A') != 'N/A':
            try:
                campos['energia_compensada_num'] = int(campos['energia_compensada'])
            except:
                campos['energia_compensada_num'] = 0
        else:
            campos['energia_compensada_num'] = 0
        
        # Saldo de créditos (energia, não monetário)
        campos['saldo_creditos_num'] = converter_valor_monetario(campos.get('saldo_creditos'))
        
        # Calcular total compensação
        campos['total_compensacao'] = campos['compensacao_tusd_num'] + campos['compensacao_te_num']
        
        # Identificar se tem sistema fotovoltaico (CRITÉRIO MELHORADO)
        tem_fotovoltaico = (campos['energia_injetada_num'] > 0 or 
                           campos['energia_compensada_num'] > 0 or 
                           campos['total_compensacao'] > 0 or
                           campos['saldo_creditos_num'] > 0)
        campos['sistema_fotovoltaico'] = 'Sim' if tem_fotovoltaico else 'Não'
        
        # Calcular economia e valor integral (conforme planilha ENEL)
        if tem_fotovoltaico and campos['valor_total_num']:
            # Valor integral sem fotovoltaico = valor atual + compensações
            campos['valor_integral_sem_fv'] = campos['valor_total_num'] + campos['total_compensacao']
            
            # Percentual de economia = (compensação / valor_integral) * 100
            if campos['valor_integral_sem_fv'] > 0:
                campos['percentual_economia_fv'] = (campos['total_compensacao'] / campos['valor_integral_sem_fv']) * 100
            else:
                campos['percentual_economia_fv'] = 0.0
        else:
            campos['valor_integral_sem_fv'] = campos.get('valor_total_num', 0.0)
            campos['percentual_economia_fv'] = 0.0
        
        return campos
    
    def processar_todos_pdfs(self) -> Dict:
        """
        Processar todos os PDFs no diretório
        
        Returns:
            Dict: Resultado do processamento
        """
        try:
            if not os.path.exists(self.pdfs_dir):
                return {
                    "status": "erro",
                    "mensagem": "Diretório de PDFs não encontrado"
                }
            
            arquivos_pdf = [f for f in os.listdir(self.pdfs_dir) if f.lower().endswith('.pdf')]
            
            if not arquivos_pdf:
                return {
                    "status": "sucesso",
                    "mensagem": "Nenhum PDF encontrado para processar",
                    "pdfs_processados": 0
                }
            
            pdfs_desprotegidos = 0
            dados_extraidos = 0
            self.dados_extraidos = []
            
            print(f"🔄 Processando {len(arquivos_pdf)} PDFs...")
            
            for i, arquivo in enumerate(arquivos_pdf, 1):
                caminho_completo = os.path.join(self.pdfs_dir, arquivo)
                
                print(f"📄 Processando {i}/{len(arquivos_pdf)}: {arquivo}")
                
                # 1. Remover proteção
                if self.remover_protecao_pdf(caminho_completo):
                    pdfs_desprotegidos += 1
                
                # 2. Extrair dados
                dados = self.extrair_dados_fatura(caminho_completo)
                if dados:
                    self.dados_extraidos.append(dados)
                    dados_extraidos += 1
            
            resultado = {
                "status": "sucesso",
                "mensagem": "Processamento de PDFs finalizado",
                "pdfs_encontrados": len(arquivos_pdf),
                "pdfs_desprotegidos": pdfs_desprotegidos,
                "dados_extraidos": dados_extraidos,
                "timestamp": datetime.now().isoformat()
            }
            
            print(f"✅ Processamento concluído:")
            print(f"   📄 PDFs: {len(arquivos_pdf)}")
            print(f"   🔓 Desprotegidos: {pdfs_desprotegidos}")
            print(f"   📊 Dados extraídos: {dados_extraidos}")
            
            return resultado
            
        except Exception as e:
            print(f"❌ Erro no processamento de PDFs: {e}")
            return {
                "status": "erro",
                "mensagem": str(e),
                "timestamp": datetime.now().isoformat()
            }
    
    def gerar_planilha_consolidada(self) -> Optional[str]:
        """
        Gerar planilha Excel consolidada SEM PANDAS
        
        Returns:
            str: Caminho da planilha gerada ou None se falhar
        """
        # Tentar versão sem pandas primeiro (Render compatible)
        resultado_sem_pandas = self.gerar_planilha_consolidada_sem_pandas()
        if resultado_sem_pandas:
            return resultado_sem_pandas
        
        # Fallback para versão com pandas (ambiente local)
        if PANDAS_AVAILABLE:
            return self._gerar_planilha_consolidada_com_pandas()
        else:
            print("⚠️ Não foi possível gerar planilha (pandas não disponível)")
            return None
        
        if not self.dados_extraidos:
            print("⚠️ Nenhum dado disponível para gerar planilha")
            return None
        
        try:
            # Preparar dados para DataFrame
            dados_planilha = []
            
            for dados in self.dados_extraidos:
                linha = {
                    'Arquivo': dados.get('arquivo', 'N/A'),
                    'Instalação': dados.get('numero_instalacao', 'N/A'),
                    'Nota Fiscal': dados.get('nota_fiscal', 'N/A'),
                    'Vencimento': dados.get('data_vencimento', 'N/A'),
                    'Competência': dados.get('competencia', 'N/A'),
                    'Valor Total': dados.get('valor_total', 'N/A'),
                    'Valor Numérico': self._validar_valor_financeiro(dados.get('valor_total_num')),
                    'Consumo kWh': dados.get('consumo_kwh', 'N/A'),
                    'Consumo Numérico': dados.get('consumo_kwh_num', 0),
                    'Data Processamento': dados.get('data_processamento', 'N/A')
                }
                dados_planilha.append(linha)
            
            # Criar DataFrame
            df = pd.DataFrame(dados_planilha)
            
            # Nome do arquivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nome_arquivo = f"faturas_enel_{timestamp}.xlsx"
            caminho_planilha = os.path.join(self.planilhas_dir, nome_arquivo)
            
            # Salvar planilha com formatação
            with pd.ExcelWriter(caminho_planilha, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Faturas ENEL', index=False)
                
                # Formatação básica
                worksheet = writer.sheets['Faturas ENEL']
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"📊 Planilha gerada: {nome_arquivo}")
            return caminho_planilha
            
        except Exception as e:
            print(f"❌ Erro gerando planilha: {e}")
            return None
    
    def gerar_planilha_consolidada_sem_pandas(self) -> Optional[str]:
        """
        Gerar planilha Excel SEM pandas (Render compatible)
        Baseado no padrão BRK que funciona perfeitamente
        
        Returns:
            str: Caminho da planilha gerada ou None se falhar
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime
        except ImportError:
            print("⚠️ openpyxl não disponível para gerar planilha")
            return None
        
        if not self.dados_extraidos:
            print("⚠️ Nenhum dado disponível para gerar planilha")
            return None
        
        try:
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Faturas ENEL"
            
            # Cabeçalhos
            cabecalhos = [
                'Arquivo', 'Instalação', 'Nota Fiscal', 'Vencimento', 
                'Competência', 'Valor Total', 'Valor Numérico', 
                'Consumo kWh', 'Consumo Numérico', 'Data Processamento'
            ]
            
            # Escrever cabeçalhos
            for col, cabecalho in enumerate(cabecalhos, 1):
                cell = ws.cell(row=1, column=col, value=cabecalho)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Escrever dados
            for row, dados in enumerate(self.dados_extraidos, 2):
                valores = [
                    dados.get('arquivo', 'N/A'),
                    dados.get('numero_instalacao', 'N/A'),
                    dados.get('nota_fiscal', 'N/A'),
                    dados.get('data_vencimento', 'N/A'),
                    dados.get('competencia', 'N/A'),
                    dados.get('valor_total', 'N/A'),
                    self._validar_valor_financeiro(dados.get('valor_total_num')),
                    dados.get('consumo_kwh', 'N/A'),
                    dados.get('consumo_kwh_num', 0),
                    dados.get('data_processamento', 'N/A')
                ]
                
                for col, valor in enumerate(valores, 1):
                    ws.cell(row=row, column=col, value=valor)
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Salvar planilha
            os.makedirs(self.planilhas_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"ENEL_Consolidado_{timestamp}.xlsx"
            caminho_planilha = os.path.join(self.planilhas_dir, nome_arquivo)
            
            wb.save(caminho_planilha)
            
            print(f"📈 Planilha gerada SEM pandas: {nome_arquivo}")
            print(f"   📄 Registros: {len(self.dados_extraidos)}")
            return caminho_planilha
            
        except Exception as e:
            print(f"❌ Erro gerando planilha sem pandas: {e}")
            return None
    
    def _gerar_planilha_consolidada_com_pandas(self) -> Optional[str]:
        """
        Versão original com pandas (fallback para ambiente local)
        """
        if not PANDAS_AVAILABLE:
            return None
        
        if not self.dados_extraidos:
            print("⚠️ Nenhum dado disponível para gerar planilha")
            return None
        
        try:
            # Preparar dados para DataFrame
            dados_planilha = []
            
            for dados in self.dados_extraidos:
                linha = {
                    'Arquivo': dados.get('arquivo', 'N/A'),
                    'Instalação': dados.get('numero_instalacao', 'N/A'),
                    'Nota Fiscal': dados.get('nota_fiscal', 'N/A'),
                    'Vencimento': dados.get('data_vencimento', 'N/A'),
                    'Competência': dados.get('competencia', 'N/A'),
                    'Valor Total': dados.get('valor_total', 'N/A'),
                    'Valor Numérico': self._validar_valor_financeiro(dados.get('valor_total_num')),
                    'Consumo kWh': dados.get('consumo_kwh', 'N/A'),
                    'Consumo Numérico': dados.get('consumo_kwh_num', 0),
                    'Data Processamento': dados.get('data_processamento', 'N/A')
                }
                dados_planilha.append(linha)
            
            # Criar DataFrame
            df = pd.DataFrame(dados_planilha)
            
            # Salvar planilha
            os.makedirs(self.planilhas_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"ENEL_Consolidado_{timestamp}.xlsx"
            caminho_planilha = os.path.join(self.planilhas_dir, nome_arquivo)
            
            # Escrever Excel com formatação
            with pd.ExcelWriter(caminho_planilha, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Faturas ENEL', index=False)
                
                # Formatar planilha
                workbook = writer.book
                worksheet = writer.sheets['Faturas ENEL']
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            print(f"📈 Planilha gerada COM pandas: {nome_arquivo}")
            return caminho_planilha
            
        except Exception as e:
            print(f"❌ Erro gerando planilha com pandas: {e}")
            return None
    
    def _validar_valor_financeiro(self, valor):
        """
        Validar valor financeiro sem fallbacks
        
        Args:
            valor: Valor a ser validado
            
        Returns:
            float ou str: Valor válido ou marcação de erro
        """
        if valor is None:
            return "ERRO_EXTRAÇÃO"
        
        if isinstance(valor, (int, float)) and valor >= 0:
            return valor
        
        return "ERRO_EXTRAÇÃO"
    
    def obter_estatisticas(self) -> Dict:
        """
        Obter estatísticas do processamento
        
        Returns:
            Dict: Estatísticas atuais
        """
        try:
            stats = {
                "sistema": "ENEL PDF Processor",
                "senha_configurada": bool(self.senha_padrao),
                "pdfs_dir": self.pdfs_dir,
                "planilhas_dir": self.planilhas_dir,
                "dados_em_cache": len(self.dados_extraidos),
                "timestamp": datetime.now().isoformat()
            }
            
            # Contar arquivos
            if os.path.exists(self.pdfs_dir):
                pdfs = [f for f in os.listdir(self.pdfs_dir) if f.lower().endswith('.pdf')]
                stats["pdfs_disponveis"] = len(pdfs)
            else:
                stats["pdfs_disponveis"] = 0
            
            if os.path.exists(self.planilhas_dir):
                planilhas = [f for f in os.listdir(self.planilhas_dir) if f.lower().endswith(('.xlsx', '.xls'))]
                stats["planilhas_geradas"] = len(planilhas)
            else:
                stats["planilhas_geradas"] = 0
            
            return stats
            
        except Exception as e:
            return {
                "erro": str(e),
                "timestamp": datetime.now().isoformat()
            }